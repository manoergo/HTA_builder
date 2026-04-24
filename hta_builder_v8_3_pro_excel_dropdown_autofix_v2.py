import re
from typing import Dict, List, Tuple, Set, Optional

import pandas as pd
import streamlit as st
from graphviz import Digraph
from graphviz.backend import ExecutableNotFound
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter

st.set_page_config(page_title="HTA Builder v8.3 Pro Excel", layout="wide")

CODE_PATTERN = re.compile(r"^\d+(?:\.\d+)*$")

# =========================
# Helpers
# =========================
def normalize_hierarchy_code(val):
    """
    Corrige automáticamente problemas típicos de Excel:
    1.0 -> 1
    0.0 -> 0
    1.3.0 -> 1.3
    mantiene 1.3.2 igual
    y convierte None/nan/celdas vacías en vacío.
    """
    if val is None:
        return ""
    try:
        if pd.isna(val):
            return ""
    except Exception:
        pass

    s = str(val).strip()
    if s.lower() in ["", "none", "nan", "nat", "null"]:
        return ""

    try:
        f = float(s)
        if f.is_integer():
            return str(int(f))
    except Exception:
        pass

    parts = s.split(".")
    cleaned = []
    for p in parts:
        p = str(p).strip()
        if p.lower() in ["", "none", "nan", "nat", "null"]:
            continue
        try:
            pf = float(p)
            if pf.is_integer():
                p = str(int(pf))
        except Exception:
            pass
        cleaned.append(p)

    while len(cleaned) > 1 and cleaned[-1] == "0":
        cleaned.pop()

    return ".".join(cleaned)

def natural_code_key(code: str) -> Tuple[int, ...]:
    """
    Ordena códigos jerárquicos de forma segura.
    Evita errores cuando Excel trae filas vacías, None o valores raros.
    """
    code = normalize_hierarchy_code(code)
    if not code:
        return (999999,)
    parts = []
    for x in str(code).split("."):
        if x == "":
            continue
        try:
            parts.append(int(float(x)))
        except Exception:
            parts.append(999999)
    return tuple(parts) if parts else (999999,)

def infer_parent(code: str) -> str:
    code = normalize_hierarchy_code(code)
    if not code:
        return ""
    parts = str(code).split(".")
    return "" if len(parts) == 1 else ".".join(parts[:-1])

def level_from_code(code: str) -> int:
    code = normalize_hierarchy_code(code)
    if not code:
        return 0
    return len(str(code).split(".")) - 1

def clean_text(x) -> str:
    if x is None:
        return ""
    try:
        if pd.isna(x):
            return ""
    except Exception:
        pass
    s = str(x).strip()
    if s.lower() in ["none", "nan", "nat", "null"]:
        return ""
    return s

def wrap_lines(text: str, width: int = 26) -> str:
    words = str(text).split()
    if not words:
        return ""
    lines = []
    line = words[0]
    for word in words[1:]:
        if len(line) + 1 + len(word) <= width:
            line += " " + word
        else:
            lines.append(line)
            line = word
    lines.append(line)
    return "\n".join(lines)

def to_csv_bytes(df: pd.DataFrame) -> bytes:
    return df.to_csv(index=False).encode("utf-8")

def make_excel_template_bytes(df: pd.DataFrame) -> bytes:
    """Crea una plantilla Excel con listas desplegables para evitar errores."""
    from io import BytesIO

    buffer = BytesIO()
    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="HTA")
        wb = writer.book
        ws = wb["HTA"]

        # Estilo de encabezados
        header_fill = PatternFill("solid", fgColor="1F4E79")
        header_font = Font(color="FFFFFF", bold=True)
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        # Congelar encabezado
        ws.freeze_panes = "A2"

        # Anchos de columna
        widths = {
            "code": 12,
            "label": 34,
            "parent": 12,
            "plan": 36,
            "type": 18,
            "notes": 42,
            "riesgo": 16,
            "esfuerzo": 18,
            "frecuencia": 18,
            "duracion": 18,
            "error": 34,
            "consecuencia": 34,
            "error_type": 20,
            "error_description": 38,
            "error_probability": 20,
            "error_severity": 20,
            "error_recovery": 18,
            "postura_forzada": 18,
            "fuerza": 14,
            "repeticion": 16,
            "contacto_estres": 20,
            "demanda_visual": 20,
            "carga_mental": 18,
            "ambiente_adverso": 20,
        }
        for idx, col_name in enumerate(df.columns, start=1):
            ws.column_dimensions[get_column_letter(idx)].width = widths.get(col_name, 18)

        # Ajuste de texto en todo el rango inicial
        for row in ws.iter_rows(min_row=2, max_row=300, min_col=1, max_col=len(df.columns)):
            for cell in row:
                cell.alignment = Alignment(vertical="top", wrap_text=True)

        # Hoja oculta con listas
        list_ws = wb.create_sheet("listas")
        list_ws.sheet_state = "hidden"

        lists = {
            "type": ["", "motora", "cognitiva", "perceptiva", "ambiental", "decisión"],
            "riesgo": ["", "muy bajo", "bajo", "medio", "alto", "crítico"],
            "esfuerzo": ["", "muy bajo", "bajo", "moderado", "alto", "muy alto"],
            "frecuencia": ["", "única", "ocasional", "repetitiva", "frecuente", "constante"],
            "duracion": ["", "muy corta", "corta", "media", "prolongada"],
            "error_type": ["", "acción", "comprobación", "selección", "información", "búsqueda", "decisión", "omisión"],
            "error_probability": ["", "baja", "media", "alta"],
            "error_severity": ["", "leve", "moderado", "severo", "crítico"],
            "error_recovery": ["", "sí", "parcial", "no"],
            "si_no": ["", "sí", "no"],
        }

        # Escribir listas en hoja oculta
        named_ranges = {}
        for col_idx, (list_name, values) in enumerate(lists.items(), start=1):
            for row_idx, value in enumerate(values, start=1):
                list_ws.cell(row=row_idx, column=col_idx, value=value)
            col_letter = get_column_letter(col_idx)
            named_ranges[list_name] = f"listas!${col_letter}$1:${col_letter}${len(values)}"

        # Mapeo de columnas a listas
        validation_map = {
            "type": "type",
            "riesgo": "riesgo",
            "esfuerzo": "esfuerzo",
            "frecuencia": "frecuencia",
            "duracion": "duracion",
            "error_type": "error_type",
            "error_probability": "error_probability",
            "error_severity": "error_severity",
            "error_recovery": "error_recovery",
            "postura_forzada": "si_no",
            "fuerza": "si_no",
            "repeticion": "si_no",
            "contacto_estres": "si_no",
            "demanda_visual": "si_no",
            "carga_mental": "si_no",
            "ambiente_adverso": "si_no",
        }

        # Crear validaciones. Notas, label, plan, error, consecuencia y descripción de error quedan libres.
        max_rows = 300
        for col_idx, col_name in enumerate(df.columns, start=1):
            if col_name in validation_map:
                list_key = validation_map[col_name]
                formula = f"={named_ranges[list_key]}"
                dv = DataValidation(
                    type="list",
                    formula1=formula,
                    allow_blank=True,
                    showErrorMessage=True,
                    errorTitle="Valor no válido",
                    error="Selecciona una opción de la lista desplegable.",
                )
                ws.add_data_validation(dv)
                col_letter = get_column_letter(col_idx)
                dv.add(f"{col_letter}2:{col_letter}{max_rows}")

        # Nota explicativa
        help_ws = wb.create_sheet("instrucciones")
        help_ws["A1"] = "Instrucciones para completar la plantilla HTA"
        help_ws["A1"].font = Font(bold=True, size=14)
        help_ws["A3"] = "Columnas obligatorias: code y label."
        help_ws["A4"] = "Las columnas con listas desplegables deben completarse seleccionando una alternativa."
        help_ws["A5"] = "Las columnas label, plan, notes, error, consecuencia y error_description son libres."
        help_ws["A6"] = "Puedes dejar columnas vacías si no aplican."
        help_ws["A7"] = "Si parent queda vacío, la app lo infiere desde el código. Ejemplo: 2.1 depende de 2."
        help_ws.column_dimensions["A"].width = 120

    buffer.seek(0)
    return buffer.getvalue()

def safe_render(dot: Digraph, fmt: str = "svg"):
    try:
        return dot.pipe(format=fmt), None
    except ExecutableNotFound:
        return None, "No se encontró el ejecutable 'dot' de Graphviz."
    except Exception as e:
        return None, str(e)

# =========================
# Maps
# =========================
RIESGO_MAP = {"": 0, "muy bajo": 1, "bajo": 2, "medio": 3, "alto": 4, "crítico": 5, "critico": 5}
ESFUERZO_MAP = {"": 0, "muy bajo": 1, "bajo": 2, "moderado": 3, "alto": 4, "muy alto": 5}
FRECUENCIA_MAP = {"": 0, "única": 1, "ocasional": 2, "repetitiva": 3, "frecuente": 4, "constante": 5}
DURACION_MAP = {"": 0, "muy corta": 1, "corta": 2, "media": 3, "prolongada": 4}
ERR_PROB_MAP = {"": 0, "baja": 1, "media": 2, "alta": 3}
ERR_SEV_MAP = {"": 0, "leve": 1, "moderado": 2, "severo": 3, "crítico": 4, "critico": 4}

FACTOR_COLS = [
    "postura_forzada",
    "fuerza",
    "repeticion",
    "contacto_estres",
    "demanda_visual",
    "carga_mental",
    "ambiente_adverso",
]

BASE_COLS = [
    "code", "label", "parent", "plan", "type", "notes",
    "riesgo", "esfuerzo", "frecuencia", "duracion",
    "error", "consecuencia",
    "error_type", "error_description", "error_probability",
    "error_severity", "error_recovery",
] + FACTOR_COLS

# =========================
# Data processing
# =========================
def preprocess_df(df: pd.DataFrame) -> pd.DataFrame:
    df = df.copy()
    for col in BASE_COLS:
        if col not in df.columns:
            df[col] = ""

    for col in BASE_COLS:
        df[col] = df[col].apply(clean_text)

    # Normalizar códigos jerárquicos provenientes de Excel
    df["code"] = df["code"].apply(normalize_hierarchy_code)
    df["parent"] = df["parent"].apply(normalize_hierarchy_code)

    # Eliminar filas completamente vacías o filas sin code y sin label.
    # Esto evita errores cuando Excel trae filas con None debajo de la tabla.
    df = df[~((df["code"] == "") & (df["label"] == ""))].copy()

    df["parent"] = df.apply(lambda r: r["parent"] if r["parent"] else infer_parent(r["code"]), axis=1)
    df["level"] = df["code"].apply(level_from_code)
    df = df.sort_values(by="code", key=lambda s: s.map(natural_code_key)).reset_index(drop=True)

    df["riesgo_score_manual"] = df["riesgo"].str.lower().map(RIESGO_MAP).fillna(0).astype(int)
    df["esfuerzo_score"] = df["esfuerzo"].str.lower().map(ESFUERZO_MAP).fillna(0).astype(int)
    df["frecuencia_score"] = df["frecuencia"].str.lower().map(FRECUENCIA_MAP).fillna(0).astype(int)
    df["duracion_score"] = df["duracion"].str.lower().map(DURACION_MAP).fillna(0).astype(int)
    df["error_prob_score"] = df["error_probability"].str.lower().map(ERR_PROB_MAP).fillna(0).astype(int)
    df["error_sev_score"] = df["error_severity"].str.lower().map(ERR_SEV_MAP).fillna(0).astype(int)

    factor_sum = []
    for _, row in df.iterrows():
        factor_sum.append(sum(1 for c in FACTOR_COLS if clean_text(row[c]).lower() == "sí"))
    df["factores_presentes"] = factor_sum

    auto_raw = (
        0.35 * df["esfuerzo_score"] +
        0.25 * df["frecuencia_score"] +
        0.20 * df["duracion_score"] +
        0.20 * df["factores_presentes"]
    )
    df["riesgo_auto_score"] = auto_raw.round(2)

    def auto_label(x: float) -> str:
        if x <= 1.0:
            return "muy bajo"
        if x <= 2.0:
            return "bajo"
        if x <= 3.0:
            return "medio"
        if x <= 4.0:
            return "alto"
        return "crítico"

    df["riesgo_auto"] = df["riesgo_auto_score"].apply(auto_label)
    df["sherpa_priority"] = df["error_prob_score"] * df["error_sev_score"]
    df["sherpa_priority_adjusted"] = df["sherpa_priority"] + df["factores_presentes"]
    return df

def validate_dataframe(df: pd.DataFrame) -> List[str]:
    errors = []
    missing = {"code", "label"} - set(df.columns)
    if missing:
        errors.append(f"Faltan columnas requeridas: {', '.join(sorted(missing))}.")
        return errors

    codes = set()
    seen = set()
    for i, row in df.iterrows():
        code = clean_text(row["code"])
        label = clean_text(row["label"])

        if not code:
            errors.append(f"Fila {i+1}: el código está vacío.")
            continue
        if not CODE_PATTERN.match(code):
            errors.append(f"Fila {i+1}: código '{code}' inválido. Usa 0, 1, 1.1, 1.2.3.")
        if code in seen:
            errors.append(f"Fila {i+1}: código duplicado '{code}'.")
        seen.add(code)

        if not label:
            errors.append(f"Fila {i+1}: la descripción de '{code}' está vacía.")
        codes.add(code)

    for i, row in df.iterrows():
        code = clean_text(row["code"])
        parent = clean_text(row.get("parent", "")) or infer_parent(code)
        if parent and parent not in codes:
            errors.append(f"Fila {i+1}: el padre '{parent}' de '{code}' no existe.")
    return errors

# =========================
# Tree helpers
# =========================
def build_children_map(df: pd.DataFrame) -> Dict[str, List[str]]:
    children: Dict[str, List[str]] = {}
    for _, row in df.iterrows():
        children.setdefault(row["parent"], []).append(row["code"])
    for k in children:
        children[k] = sorted(children[k], key=natural_code_key)
    return children

def descendants(root: str, children_map: Dict[str, List[str]]) -> Set[str]:
    result = {root}
    stack = [root]
    while stack:
        cur = stack.pop()
        for child in children_map.get(cur, []):
            if child not in result:
                result.add(child)
                stack.append(child)
    return result

def subtree_df(df: pd.DataFrame, root_code: str) -> pd.DataFrame:
    cmap = build_children_map(df)
    keep = descendants(root_code, cmap)
    out = df[df["code"].isin(keep)].copy()
    return out.sort_values(by="code", key=lambda s: s.map(natural_code_key)).reset_index(drop=True)

def filter_to_max_level(df: pd.DataFrame, max_level: Optional[int]) -> pd.DataFrame:
    if max_level is None:
        return df.copy()
    return df[df["level"] <= max_level].copy().reset_index(drop=True)

# =========================
# Graph
# =========================
def node_fill(row: pd.Series, color_mode: str, paper_mode: bool) -> str:
    if paper_mode:
        gray_risk = {"muy bajo": "#ffffff", "bajo": "#efefef", "medio": "#dddddd", "alto": "#c0c0c0", "crítico": "#9d9d9d", "critico": "#9d9d9d"}
        if color_mode == "Riesgo automático":
            return gray_risk.get(clean_text(row["riesgo_auto"]).lower(), "#ffffff")
        return "#ffffff"

    risk_map = {"muy bajo": "#e5e7eb", "bajo": "#dcfce7", "medio": "#fef3c7", "alto": "#fed7aa", "crítico": "#fecaca", "critico": "#fecaca"}
    sherpa_map = {"leve": "#e5e7eb", "moderado": "#fde68a", "severo": "#fdba74", "crítico": "#fecaca", "critico": "#fecaca"}
    type_map = {"cognitiva": "#dbeafe", "motora": "#dcfce7", "perceptiva": "#fef3c7", "ambiental": "#f3e8ff", "decisión": "#fee2e2", "decision": "#fee2e2"}
    level_map = {0: "#f3f4f6", 1: "#e5f3ff", 2: "#eefce8", 3: "#fff4df", 4: "#f5ecff", 5: "#ffe8ef"}
    factor_map = {0: "#ffffff", 1: "#e0f2fe", 2: "#fef3c7", 3: "#fed7aa", 4: "#fecaca", 5: "#fecaca", 6: "#fecaca", 7: "#fecaca"}

    if color_mode == "Riesgo automático":
        return risk_map.get(clean_text(row["riesgo_auto"]).lower(), "#ffffff")
    if color_mode == "Riesgo manual":
        return risk_map.get(clean_text(row["riesgo"]).lower(), "#ffffff")
    if color_mode == "SHERPA":
        return sherpa_map.get(clean_text(row["error_severity"]).lower(), "#ffffff")
    if color_mode == "Tipo de tarea":
        return type_map.get(clean_text(row["type"]).lower(), "#ffffff")
    if color_mode == "Factores ergonómicos":
        return factor_map.get(int(row["factores_presentes"]), "#ffffff")
    return level_map.get(int(row["level"]), "#ffffff")

def make_node_label(
    row: pd.Series,
    modo_nodo: str,
    show_risk: bool,
    risk_source: str,
    show_effort: bool,
    show_frequency: bool,
    show_duration: bool,
    show_sherpa: bool,
    show_factors: bool,
    show_notes: bool,
    wrap_width: int,
) -> str:
    lines = [str(row["code"]), wrap_lines(row["label"], wrap_width)]

    if show_risk:
        if risk_source == "Automático" and row["riesgo_auto"]:
            lines.append(f'Riesgo: {row["riesgo_auto"]}')
        elif risk_source == "Manual" and row["riesgo"]:
            lines.append(f'Riesgo: {row["riesgo"]}')
        elif row["riesgo_auto"]:
            lines.append(f'Riesgo: {row["riesgo_auto"]}')

    if modo_nodo == "Detallado":
        if show_effort and row["esfuerzo"]:
            lines.append(f'Esfuerzo: {row["esfuerzo"]}')
        if show_frequency and row["frecuencia"]:
            lines.append(f'Freq: {row["frecuencia"]}')
        if show_duration and row["duracion"]:
            lines.append(f'Dur: {row["duracion"]}')
        if show_sherpa:
            if row["error_severity"]:
                lines.append(f'SHERPA: {row["error_severity"]}')
            if row["sherpa_priority_adjusted"] > 0:
                lines.append(f'P: {int(row["sherpa_priority_adjusted"])}')
        if show_factors and row["factores_presentes"] > 0:
            lines.append(f'Factores: {int(row["factores_presentes"])}')
        if show_notes and row["notes"]:
            lines.append(wrap_lines(f'Obs: {row["notes"]}', wrap_width))

    return "\n".join(lines)

def build_hta_graph(
    df: pd.DataFrame,
    color_mode: str,
    modo_nodo: str,
    show_risk: bool,
    risk_source: str,
    show_effort: bool,
    show_frequency: bool,
    show_duration: bool,
    show_sherpa: bool,
    show_factors: bool,
    show_notes: bool,
    show_plans: bool,
    paper_mode: bool,
    font_size: int = 11,
    plan_font_size: int = 10,
    wrap_width: int = 18,
    node_margin: str = "0.10,0.08",
    nodesep: float = 0.5,
    ranksep: float = 0.8,
) -> Digraph:
    dot = Digraph("HTA")
    dot.attr(rankdir="TB", splines="polyline", bgcolor="white",
             nodesep=str(nodesep), ranksep=str(ranksep), pad="0.6", margin="0.2")
    dot.attr("node", shape="box", style="rounded,filled" if not paper_mode else "filled",
             color="black", penwidth="1.2" if paper_mode else "1.4", fontname="Arial")
    dot.attr("edge", color="black", penwidth="1.0" if paper_mode else "1.2", arrowsize="0.0")

    available_codes = set(df["code"])
    for _, row in df.iterrows():
        label = make_node_label(
            row, modo_nodo, show_risk, risk_source, show_effort, show_frequency,
            show_duration, show_sherpa, show_factors, show_notes, wrap_width
        )
        dot.node(
            row["code"],
            label=label,
            fillcolor=node_fill(row, color_mode, paper_mode),
            fontsize=str(font_size),
            margin=node_margin,
        )

    for _, row in df.iterrows():
        if row["parent"] and row["parent"] in available_codes:
            dot.edge(row["parent"], row["code"])

    parent_groups: Dict[str, List[str]] = {}
    for _, row in df.iterrows():
        parent_groups.setdefault(row["parent"], []).append(row["code"])
    for parent, children in parent_groups.items():
        if parent and len(children) > 1:
            with dot.subgraph() as s:
                s.attr(rank="same")
                for child in sorted(children, key=natural_code_key):
                    s.node(child)

    if show_plans:
        for _, row in df.iterrows():
            if row["plan"]:
                plan_id = f'plan_{row["code"].replace(".", "_")}'
                plan_text = f'Plan {row["code"]}\n{wrap_lines(row["plan"], 28)}'
                dot.node(plan_id, label=plan_text, shape="note", style="filled",
                         fillcolor="#ffffff", color="#666666", penwidth="0.8",
                         fontsize=str(plan_font_size), fontname="Arial")
                dot.edge(row["code"], plan_id, style="invis", weight="25")
                with dot.subgraph() as s:
                    s.attr(rank="same")
                    s.node(row["code"])
                    s.node(plan_id)
    return dot

def build_hta_graph_export(**kwargs) -> Digraph:
    dpi = kwargs.pop("dpi", 300)
    dot = build_hta_graph(**kwargs)
    dot.attr(dpi=str(dpi))
    return dot

# =========================
# Example data
# =========================
def example_dataframe() -> pd.DataFrame:
    rows = [
        ["0","Ir al baño en hospital sin asistencia","","hacer 1, 2 y 3 en secuencia","","Objetivo principal del análisis","medio","","","","","","","","","","","","","","","","",""],
        ["1","Decidir ir al baño","0","hacer 1.1, 1.2 y 1.3","decisión","Proceso inicial de valoración","medio","moderado","ocasional","corta","Decisión tardía","Demora o acción insegura","decisión","Decide actuar sin ayuda","media","moderado","sí","no","no","no","no","no","sí","no"],
        ["1.1","Percibir necesidad fisiológica","1","","perceptiva","Puede alterarse por medicación o fatiga","bajo","bajo","ocasional","corta","","","información","No reconoce adecuadamente la necesidad","baja","leve","sí","no","no","no","no","sí","no","no"],
        ["1.2","Evaluar urgencia","1","","cognitiva","Depende de interpretación del malestar","medio","moderado","ocasional","corta","","","comprobación","Evalúa mal la urgencia","media","moderado","sí","no","no","no","no","sí","sí","no"],
        ["1.3","Evaluar capacidad para movilizarse","1","","cognitiva","Tarea crítica por posible sobreestimación","alto","moderado","ocasional","corta","Subestimar limitaciones","Riesgo de caída","comprobación","Sobreestima su capacidad física","alta","severo","no","no","no","no","no","no","sí","sí"],
        ["2","Salir de la cama sin ayuda del staff","0","hacer 2.1, 2.2 y 2.3","motora","Bloque con mayor exigencia física","alto","alto","ocasional","media","Transferencia inestable","Caída o sobrecarga","","","","","","sí","sí","no","no","no","no","sí"],
        ["2.1","Ajustar posición en la cama","2","","motora","Movimiento preparatorio","medio","moderado","ocasional","corta","","","acción","Movimiento ineficaz","media","moderado","sí","sí","sí","no","no","no","no","no"],
        ["2.2","Girar el cuerpo hacia el borde","2","","motora","Requiere coordinación y control","medio","moderado","ocasional","corta","","","acción","Giro descontrolado","media","moderado","sí","sí","sí","no","no","no","no","no"],
        ["2.3","Transferirse a bipedestación","2","","motora","Punto de máximo riesgo","crítico","muy alto","ocasional","corta","Pérdida de equilibrio","Caída","acción","Pérdida de equilibrio al ponerse de pie","alta","crítico","no","sí","sí","no","no","no","no","sí"],
        ["3","Prepararse para caminar","0","hacer 3.1 y 3.2","motora","Bloque de transición funcional","alto","alto","frecuente","media","","","","","","","","sí","sí","sí","no","no","no","sí"],
        ["3.1","Evaluar estabilidad corporal","3","","cognitiva","Verificación previa al desplazamiento","medio","moderado","ocasional","corta","","","comprobación","No identifica inestabilidad","media","moderado","sí","no","no","no","no","no","sí","no"],
        ["3.2","Iniciar marcha","3","","motora","Inicio puede ser inseguro","alto","alto","frecuente","media","","","acción","Inicio de marcha sin control","alta","severo","parcial","sí","sí","sí","no","no","no","sí"],
    ]
    return pd.DataFrame(rows, columns=BASE_COLS)

# =========================
# UI
# =========================
st.title("HTA Builder v8.3 Pro Excel")
st.caption("Versión integrada con carga de datos desde Excel usando columnas definidas.")
st.markdown(
    "**Hecho por el Dr. Carlos Manuel Escobar Galindo**  \n"
    "Doctor en Ergonomía  \n"
    "Contacto: **cescobarg@unmsm.edu.pe** | **carlos.escobar@upc.edu.pe**"
)

with st.sidebar:
    st.header("Créditos")
    st.markdown(
        "**Hecho por el Dr. Carlos Manuel Escobar Galindo**  \n"
        "Doctor en Ergonomía  \n"
        "cescobarg@unmsm.edu.pe  \n"
        "carlos.escobar@upc.edu.pe"
    )

    st.header("Módulos")
    usar_sherpa = st.toggle("Activar SHERPA", value=True)
    usar_factores = st.toggle("Activar factores ergonómicos", value=True)

    st.header("Visualización")
    color_mode = st.selectbox("Colorear nodos por", ["Riesgo automático", "Riesgo manual", "SHERPA", "Tipo de tarea", "Factores ergonómicos", "Nivel"], index=0)
    modo_nodo = st.radio("Modo de nodo", ["Resumido", "Detallado"], index=0)
    show_risk = st.toggle("Mostrar riesgo en nodos", value=False)
    risk_source = st.selectbox("Tipo de riesgo en nodos", ["Automático", "Manual"], index=0)
    show_effort = st.toggle("Mostrar esfuerzo en nodos", value=False)
    show_frequency = st.toggle("Mostrar frecuencia en nodos", value=False)
    show_duration = st.toggle("Mostrar duración en nodos", value=False)
    show_sherpa = st.toggle("Mostrar datos SHERPA en nodos", value=False)
    show_factors = st.toggle("Mostrar factores en nodos", value=False)
    show_notes = st.toggle("Mostrar notas/observaciones en nodos", value=False)
    show_plans = st.toggle("Mostrar planes", value=True)
    view_mode = st.radio("Modo de visualización", ["HTA completo", "Dividir por tarea principal"], index=1)
    max_level_option = st.selectbox("Mostrar hasta nivel", ["Todos", "0", "1", "2", "3", "4", "5", "6"], index=4)

    st.header("Ajuste visual")
    font_size = st.slider("Tamaño de texto del nodo", 8, 24, 11)
    plan_font_size = st.slider("Tamaño de texto del plan", 8, 20, 10)
    wrap_width = st.slider("Ancho de texto por línea", 16, 40, 18)
    margin_x = st.slider("Padding horizontal del nodo", 8, 30, 10)
    margin_y = st.slider("Padding vertical del nodo", 6, 20, 8)
    nodesep = st.slider("Separación horizontal entre nodos", 0.2, 1.5, 0.50, 0.05)
    ranksep = st.slider("Separación vertical entre niveles", 0.2, 1.8, 0.80, 0.05)

    st.header("Exportación")
    paper_mode = st.toggle("Modo paper (blanco y negro)", value=False)
    paper_dpi = st.selectbox("Resolución de exportación", [150, 300, 600], index=1)
    export_png = st.toggle("Habilitar PNG", value=True)
    export_jpg = st.toggle("Habilitar JPG", value=True)
    export_svg = st.toggle("Habilitar SVG", value=True)
    modo_paper = st.toggle("Mostrar tablas paper", value=True)

max_level = None if max_level_option == "Todos" else int(max_level_option)
node_margin = f"{margin_x/100:.2f},{margin_y/100:.2f}"

tab1, tab2, tab3 = st.tabs(["Editor manual", "Subir Excel", "Plantilla Excel"])

with tab1:
    df0 = example_dataframe()
    visible_cols = ["code","label","parent","plan","type","notes","riesgo","esfuerzo","frecuencia","duracion","error","consecuencia"]
    if usar_sherpa:
        visible_cols += ["error_type","error_description","error_probability","error_severity","error_recovery"]
    if usar_factores:
        visible_cols += FACTOR_COLS
    editor_df = df0[visible_cols].copy()

    column_config = {
        "code": st.column_config.TextColumn("code"),
        "label": st.column_config.TextColumn("label", width="large"),
        "parent": st.column_config.TextColumn("parent"),
        "plan": st.column_config.TextColumn("plan", width="large"),
        "type": st.column_config.SelectboxColumn("type", options=["", "motora", "cognitiva", "perceptiva", "ambiental", "decisión"]),
        "notes": st.column_config.TextColumn("notes", width="large"),
        "riesgo": st.column_config.SelectboxColumn("riesgo", options=["", "muy bajo", "bajo", "medio", "alto", "crítico"]),
        "esfuerzo": st.column_config.SelectboxColumn("esfuerzo", options=["", "muy bajo", "bajo", "moderado", "alto", "muy alto"]),
        "frecuencia": st.column_config.SelectboxColumn("frecuencia", options=["", "única", "ocasional", "repetitiva", "frecuente", "constante"]),
        "duracion": st.column_config.SelectboxColumn("duracion", options=["", "muy corta", "corta", "media", "prolongada"]),
        "error": st.column_config.TextColumn("error", width="large"),
        "consecuencia": st.column_config.TextColumn("consecuencia", width="large"),
    }
    if usar_sherpa:
        column_config.update({
            "error_type": st.column_config.SelectboxColumn("error_type", options=["", "acción", "comprobación", "selección", "información", "búsqueda", "decisión", "omisión"]),
            "error_description": st.column_config.TextColumn("error_description", width="large"),
            "error_probability": st.column_config.SelectboxColumn("error_probability", options=["", "baja", "media", "alta"]),
            "error_severity": st.column_config.SelectboxColumn("error_severity", options=["", "leve", "moderado", "severo", "crítico"]),
            "error_recovery": st.column_config.SelectboxColumn("error_recovery", options=["", "sí", "parcial", "no"]),
        })
    if usar_factores:
        for c in FACTOR_COLS:
            column_config[c] = st.column_config.SelectboxColumn(c, options=["", "sí", "no"])

    edited_df = st.data_editor(editor_df, num_rows="dynamic", use_container_width=True, hide_index=True, column_config=column_config, key="editor_v83")

with tab2:
    st.markdown(
        "Sube un archivo **Excel (.xlsx)** con las columnas definidas de la matriz HTA. "
        "La app leerá la primera hoja del archivo."
    )
    uploaded = st.file_uploader("Sube tu plantilla Excel HTA (.xlsx)", type=["xlsx"])
    upload_df = None
    if uploaded is not None:
        try:
            upload_df = pd.read_excel(uploaded, sheet_name=0, dtype={"code": str, "parent": str}, keep_default_na=False)

            # corregir automáticamente códigos dañados por Excel
            if "code" in upload_df.columns:
                upload_df["code"] = upload_df["code"].apply(normalize_hierarchy_code)

            if "parent" in upload_df.columns:
                upload_df["parent"] = upload_df["parent"].apply(normalize_hierarchy_code)

            st.success("Excel cargado correctamente. Se corrigieron automáticamente formatos tipo 1.0 → 1.")
            st.dataframe(upload_df, use_container_width=True)
        except Exception as e:
            st.error(f"No se pudo leer el archivo Excel: {e}")

with tab3:
    tmpl = example_dataframe()
    st.markdown("Descarga esta plantilla, edítala en Excel y luego súbela en la pestaña **Subir Excel**.")
    st.dataframe(tmpl.head(10), use_container_width=True)

    st.download_button(
        "Descargar plantilla Excel con listas desplegables (.xlsx)",
        data=make_excel_template_bytes(tmpl),
        file_name="hta_template_v8_3_pro_excel_dropdown.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        use_container_width=True,
    )

source_df = upload_df if "upload_df" in locals() and upload_df is not None else edited_df

with st.expander("Guía rápida de columnas para completar el Excel"):
    st.markdown(
        """
        **Columnas mínimas obligatorias:** `code`, `label`.

        **Columnas de estructura HTA:**  
        `code`, `label`, `parent`, `plan`, `type`, `notes`.

        **Columnas ergonómicas:**  
        `riesgo`, `esfuerzo`, `frecuencia`, `duracion`, `error`, `consecuencia`.

        **Columnas SHERPA:**  
        `error_type`, `error_description`, `error_probability`, `error_severity`, `error_recovery`.

        **Columnas de factores ergonómicos:**  
        `postura_forzada`, `fuerza`, `repeticion`, `contacto_estres`, `demanda_visual`, `carga_mental`, `ambiente_adverso`.

        Puedes dejar columnas vacías. La app completará automáticamente `parent` si el código tiene jerarquía, por ejemplo `2.1` depende de `2`.
        """
    )

def render_download_buttons(dot_export: Digraph, df_to_download: pd.DataFrame, prefix: str):
    cols = st.columns(5)
    cols[0].download_button("DOT", dot_export.source, f"{prefix}.dot", "text/plain", use_container_width=True, key=f"{prefix}_dot")
    if export_svg:
        svg_bytes, svg_error = safe_render(dot_export, "svg")
        if svg_bytes is not None:
            cols[1].download_button("SVG", svg_bytes, f"{prefix}.svg", "image/svg+xml", use_container_width=True, key=f"{prefix}_svg")
        else:
            cols[1].button("SVG no disponible", disabled=True, use_container_width=True, key=f"{prefix}_svg_na")
            if svg_error:
                st.warning(svg_error)
    if export_png:
        png_bytes, png_error = safe_render(dot_export, "png")
        if png_bytes is not None:
            cols[2].download_button("PNG", png_bytes, f"{prefix}.png", "image/png", use_container_width=True, key=f"{prefix}_png")
        else:
            cols[2].button("PNG no disponible", disabled=True, use_container_width=True, key=f"{prefix}_png_na")
            if png_error:
                st.warning(png_error)
    if export_jpg:
        jpg_bytes, jpg_error = safe_render(dot_export, "jpg")
        if jpg_bytes is not None:
            cols[3].download_button("JPG", jpg_bytes, f"{prefix}.jpg", "image/jpeg", use_container_width=True, key=f"{prefix}_jpg")
        else:
            cols[3].button("JPG no disponible", disabled=True, use_container_width=True, key=f"{prefix}_jpg_na")
            if jpg_error:
                st.warning(jpg_error)
    cols[4].download_button("CSV", to_csv_bytes(df_to_download), f"{prefix}.csv", "text/csv", use_container_width=True, key=f"{prefix}_csv")

if st.button("Generar HTA v8.3 Pro Excel", type="primary", use_container_width=True):
    df = preprocess_df(source_df)
    errors = validate_dataframe(df)

    if errors:
        st.error("Corrige estos problemas antes de generar el HTA:")
        for err in errors:
            st.write(f"- {err}")
    else:
        filtered_df = filter_to_max_level(df, max_level)

        if not usar_sherpa:
            for c in ["error_type","error_description","error_probability","error_severity","error_recovery"]:
                filtered_df[c] = ""
            filtered_df["error_prob_score"] = 0
            filtered_df["error_sev_score"] = 0
            filtered_df["sherpa_priority"] = 0
            filtered_df["sherpa_priority_adjusted"] = 0
            show_sherpa = False

        if not usar_factores:
            for c in FACTOR_COLS:
                filtered_df[c] = ""
            filtered_df["factores_presentes"] = 0
            show_factors = False

        st.success("HTA generado.")

        st.subheader("Panel de análisis del sistema")
        c1, c2, c3, c4, c5 = st.columns(5)
        total = len(filtered_df)
        altos = int(filtered_df["riesgo_auto"].isin(["alto", "crítico"]).sum())
        criticos = int(filtered_df["riesgo_auto"].isin(["crítico"]).sum())
        con_factores = int((filtered_df["factores_presentes"] > 0).sum())
        c1.metric("Total tareas", total)
        c2.metric("Alto + crítico", altos)
        c3.metric("Críticas", criticos)
        c4.metric("Con factores", con_factores)
        c5.metric("% alto riesgo", f"{(altos/total*100):.1f}%" if total else "0%")

        st.subheader("Matriz maestra")
        matrix_cols = ["code","label","parent","type","notes","riesgo","riesgo_auto","esfuerzo","frecuencia","duracion","factores_presentes","error","consecuencia"]
        if usar_sherpa:
            matrix_cols += ["error_type","error_description","error_probability","error_severity","error_recovery","sherpa_priority_adjusted"]
        st.dataframe(filtered_df[matrix_cols], use_container_width=True, hide_index=True)

        st.subheader("Heatmap de riesgo")
        def color_riesgo(val):
            colores = {"muy bajo": "#e5e7eb","bajo": "#bbf7d0","medio": "#fde68a","alto": "#fb923c","crítico": "#ef4444","critico": "#ef4444"}
            return f"background-color: {colores.get(str(val).lower(), '#ffffff')}"
        heatmap_df = filtered_df[["code","label","riesgo_auto"]].copy()
        st.dataframe(heatmap_df.style.map(color_riesgo, subset=["riesgo_auto"]), use_container_width=True, hide_index=True)

        st.subheader("Filtro de análisis")
        filtro = st.selectbox("Filtrar tareas", ["Todas", "Críticas", "Altas", "Con error SHERPA", "Con factores"])
        df_filtrado = filtered_df.copy()
        if filtro == "Críticas":
            df_filtrado = df_filtrado[df_filtrado["riesgo_auto"] == "crítico"]
        elif filtro == "Altas":
            df_filtrado = df_filtrado[df_filtrado["riesgo_auto"].isin(["alto", "crítico"])]
        elif filtro == "Con error SHERPA":
            df_filtrado = df_filtrado[df_filtrado["error_sev_score"] > 0]
        elif filtro == "Con factores":
            df_filtrado = df_filtrado[df_filtrado["factores_presentes"] > 0]
        st.dataframe(df_filtrado, use_container_width=True, hide_index=True)

        top_level = sorted(filtered_df[filtered_df["parent"] == "0"]["code"].tolist(), key=natural_code_key)

        if view_mode == "HTA completo":
            dot_view = build_hta_graph(
                filtered_df, color_mode, modo_nodo, show_risk, risk_source, show_effort,
                show_frequency, show_duration, show_sherpa and usar_sherpa,
                show_factors and usar_factores, show_notes, show_plans, paper_mode,
                font_size, plan_font_size, wrap_width, node_margin, nodesep, ranksep
            )
            st.subheader("Vista HTA")
            st.graphviz_chart(dot_view, use_container_width=True)

            dot_export = build_hta_graph_export(
                df=filtered_df, color_mode=color_mode, modo_nodo=modo_nodo,
                show_risk=show_risk, risk_source=risk_source, show_effort=show_effort,
                show_frequency=show_frequency, show_duration=show_duration,
                show_sherpa=show_sherpa and usar_sherpa,
                show_factors=show_factors and usar_factores, show_notes=show_notes,
                show_plans=show_plans, paper_mode=paper_mode, font_size=font_size,
                plan_font_size=plan_font_size, wrap_width=wrap_width,
                node_margin=node_margin, nodesep=nodesep, ranksep=ranksep, dpi=paper_dpi
            )
            render_download_buttons(dot_export, filtered_df, "hta_v8_3_pro_excel")
        else:
            st.subheader("Vista dividida por tarea principal")
            if not top_level:
                st.info("No se encontraron tareas principales hijas de 0 con el filtro actual.")
            for code in top_level:
                label = filtered_df.loc[filtered_df["code"] == code, "label"].iloc[0]
                st.markdown(f"## {code} - {label}")
                part_df = subtree_df(filtered_df, code)

                dot_view = build_hta_graph(
                    part_df, color_mode, modo_nodo, show_risk, risk_source, show_effort,
                    show_frequency, show_duration, show_sherpa and usar_sherpa,
                    show_factors and usar_factores, show_notes, show_plans, paper_mode,
                    font_size, plan_font_size, wrap_width, node_margin, nodesep, ranksep
                )
                st.graphviz_chart(dot_view, use_container_width=True)
                st.dataframe(part_df[matrix_cols], use_container_width=True, hide_index=True)

                dot_export = build_hta_graph_export(
                    df=part_df, color_mode=color_mode, modo_nodo=modo_nodo,
                    show_risk=show_risk, risk_source=risk_source, show_effort=show_effort,
                    show_frequency=show_frequency, show_duration=show_duration,
                    show_sherpa=show_sherpa and usar_sherpa,
                    show_factors=show_factors and usar_factores, show_notes=show_notes,
                    show_plans=show_plans, paper_mode=paper_mode, font_size=font_size,
                    plan_font_size=plan_font_size, wrap_width=wrap_width,
                    node_margin=node_margin, nodesep=nodesep, ranksep=ranksep, dpi=paper_dpi
                )
                render_download_buttons(dot_export, part_df, f"hta_{code.replace('.', '_')}_pro_excel")

        st.subheader("Priorización automática")
        priority_df = filtered_df[["code","label","riesgo_auto","riesgo_auto_score","factores_presentes"]].copy()
        priority_df["sherpa_prioridad"] = filtered_df["sherpa_priority_adjusted"] if usar_sherpa else 0

        def prioridad_intervencion(row):
            score = row["riesgo_auto_score"] + row["factores_presentes"] + row["sherpa_prioridad"]
            if score >= 12:
                return "intervención inmediata"
            if score >= 8:
                return "intervención prioritaria"
            if score >= 5:
                return "seguimiento"
            return "baja prioridad"

        priority_df["prioridad_intervencion"] = priority_df.apply(prioridad_intervencion, axis=1)
        st.dataframe(priority_df, use_container_width=True, hide_index=True)

        if modo_paper:
            st.subheader("Tablas para paper")
            table_hta = filtered_df[["code","label","parent","plan","type","notes"]].copy()
            st.markdown("**Tabla 1. Estructura HTA**")
            st.dataframe(table_hta, use_container_width=True, hide_index=True)

            table_ergo = filtered_df[["code","label","riesgo","riesgo_auto","esfuerzo","frecuencia","duracion","factores_presentes","error","consecuencia","notes"]].copy()
            st.markdown("**Tabla 2. Matriz ergonómica**")
            st.dataframe(table_ergo, use_container_width=True, hide_index=True)

            paper_export = {"hta_table": table_hta, "ergo_table": table_ergo}

            if usar_sherpa:
                table_sherpa = filtered_df[["code","label","error_type","error_description","error_probability","error_severity","error_recovery","sherpa_priority_adjusted","notes"]].copy()
                st.markdown("**Tabla 3. Matriz SHERPA**")
                st.dataframe(table_sherpa, use_container_width=True, hide_index=True)
                paper_export["sherpa_table"] = table_sherpa

            if usar_factores:
                table_factores = filtered_df[["code","label"] + FACTOR_COLS + ["factores_presentes","notes"]].copy()
                st.markdown("**Tabla 4. Factores de riesgo ergonómico**")
                st.dataframe(table_factores, use_container_width=True, hide_index=True)
                paper_export["factores_table"] = table_factores

            paper_csv = pd.concat(
                [d.assign(tabla=name) for name, d in paper_export.items()],
                ignore_index=True,
                sort=False,
            ).to_csv(index=False).encode("utf-8")

            st.download_button(
                "Descargar tablas paper (CSV unificado)",
                data=paper_csv,
                file_name="hta_v8_3_pro_excel_tablas_paper.csv",
                mime="text/csv",
                use_container_width=True,
            )

st.markdown("---")
st.markdown(
    "Desarrollado por **Dr. Carlos Manuel Escobar Galindo**, Doctor en Ergonomía.  \n"
    "Contacto: **cescobarg@unmsm.edu.pe** | **carlos.escobar@upc.edu.pe**"
)
